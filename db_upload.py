from firebase import Firebase
import openpyxl

path = 'Participants List DTUPD23.xlsx'

wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active

config = {

  "apiKey": "AIzaSyDvadj2Mc_vAUDYGOPJqA-gt7SiZcag_is",

  "authDomain": "dtupd2022fb.firebaseapp.com",

  "databaseURL": "https://dtupd2022fb-default-rtdb.asia-southeast1.firebasedatabase.app",

  "projectId": "dtupd2022fb",

  "storageBucket": "dtupd2022fb.appspot.com",

  "messagingSenderId": "68911700542",

  "appId": "1:68911700542:web:56b0096bfedeae50638f6f",

  "measurementId": "G-D3QNTJ4MC4"

}

firebase = Firebase(config)

db = firebase.database()

for i in range (1,194):

    num = sheet_obj.cell(row=i, column=1)
    name = sheet_obj.cell(row=i,column=2)

    toUpload = {
        "Name":name.value,
        "Check_In_7th" : 0,
        "Meal_1_7th": 0,
        "Meal_2_7th": 0,
        "Check_In_8th":0,
        "Meal_1_8th": 0,
        "Meal_2_8th": 0,
        "Check_In_9th":0,
        "Meal_1_9th": 0,
        "Meal_2_9th": 0
        }
    db.child("Participants/"+str(num.value)).set(toUpload)

# db.set(toUpload)
