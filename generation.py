import pyqrcode
import png
from pyqrcode import QRCode
import openpyxl

path = 'Participants List DTUPD23.xlsx'

wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active

# for i in range (194,197):

#     id_num = sheet_obj.cell(row=i, column=1)
#     id_name = sheet_obj.cell(row=i,column=2)

#     # print(id_num.value)
#     # print(id_name.value)

qr = pyqrcode.create(69299)
qr.png(str(69299)+'_'+'kabir'+'.png', scale=6)

# # String which represents the QR code
# s = "hello"


# # Generate QR code
# url = pyqrcode.create(s)

# # Create and save the png file naming "myqr.png"
# url.png('myqr.png', scale = 6)